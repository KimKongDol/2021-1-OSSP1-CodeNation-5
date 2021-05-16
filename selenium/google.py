from selenium import webdriver
import time
from openpyxl import Workbook
import pandas as pd

wb = Workbook()
ws = wb.create_sheet()
driver = webdriver.Chrome("/Users/iseulgi/Downloads/chromedriver")

final = {}

keyword = input('검색어를 입력하세요:')
name = keyword

#키워드 검색
URL = "https://www.google.com/search?q={}".format(keyword)
driver.get(URL)
driver.implicitly_wait(3)
searches = driver.find_elements_by_css_selector(".s75CSd.ZJjDdf.LTmh7d")
temp = []
for keyword in searches:
    try:
        result = keyword.text
        temp.append(result)
    except:
        pass
final[result] = pd.Series(temp)

keyword_list = []

# 키워드 세개를 클릭하면 실행될 함수
if (len(keyword_list) > 2):
    try:
        def keywordList():
            for i in range(len(keyword)):
                URL = "https://www.google.com/search?q={}".format(keyword)
                driver.get(URL)
                driver.implicitly_wait(3)

                searches = driver.find_elements_by_css_selector(".lst_related_srch li")

                temp = []

                for keyword in searches:
                    result = keyword.text
                    temp.append(result)

                final[keyword_list[i]] = pd.Series(temp)
    except: #예외가 발생하더라도 문제없이 다음코드를 실행할 수 있다.
        pass


df = pd.DataFrame(final)

#키워드 엑셀파일에 저장
df.to_excel(str(name) +'.xlsx')

wb = openyxl.load_workbook(str(name) +'.xlsx') #엑셀파일 열기

ws = wb.get_sheet_by_name("Sheet1") #첫번째 시트 얻기
#ws = wb.active
m_col = ws.max_column #시트의 열 갯수 읽어오기
variable_ea = m_col - 1 #변수로 네이밍할 예정

#변수 저장
for i in range(1, variable_ea): 
    j = i+1
    data = ws['B'+ j].value
    globals()['keyword'+str(i)]= data

driver.quit()

